"""
Convert Excel named ranges to text input file for FP.c program
Requires: openpyxl (install with: pip install openpyxl)
"""

import openpyxl
from openpyxl.utils import get_column_letter

def excel_to_input(excel_file, output_file):
    """
    Read data from Excel named ranges and write to text file format.
    
    Expected named ranges in Excel:
    - 'Airports': Column range with format [Code, Latitude, Longitude]
    - 'Constraints': Column range with format [Begin_Code, End_Code]
    - 'StartAirport': Single cell with start airport code (optional)
    """
    
    try:
        wb = openpyxl.load_workbook(excel_file)
    except FileNotFoundError:
        print(f"Error: Excel file '{excel_file}' not found")
        return False
    except Exception as e:
        print(f"Error opening Excel file: {e}")
        return False
    
    # Get the named ranges from the workbook
    named_ranges = wb.named_ranges
    
    airports_data = []
    constraints_data = []
    start_airport = None
    
    try:
        # Read Airports named range
        if 'Airports' in [nr.name for nr in named_ranges]:
            airports_range = wb.named_ranges['Airports']
            for cell in airports_range.destinations:
                ws = wb[cell[0]]
                for row in ws[cell[1]]:
                    code = row[0].value
                    lat = row[1].value
                    lon = row[2].value
                    if code and lat is not None and lon is not None:
                        airports_data.append((code, lat, lon))
        else:
            print("Warning: 'Airports' named range not found")
            return False
        
        # Read Constraints named range
        if 'Constraints' in [nr.name for nr in named_ranges]:
            constraints_range = wb.named_ranges['Constraints']
            for cell in constraints_range.destinations:
                ws = wb[cell[0]]
                for row in ws[cell[1]]:
                    beg = row[0].value
                    end = row[1].value
                    if beg and end:
                        constraints_data.append((beg, end))
        else:
            print("Warning: 'Constraints' named range not found")
        
        # Read StartAirport named range (optional)
        if 'StartAirport' in [nr.name for nr in named_ranges]:
            start_range = wb.named_ranges['StartAirport']
            for cell in start_range.destinations:
                ws = wb[cell[0]]
                start_airport = ws[cell[1]].value
        
    except Exception as e:
        print(f"Error reading named ranges: {e}")
        return False
    
    # Write to output file
    try:
        with open(output_file, 'w') as f:
            # Write number of airports
            f.write(f"{len(airports_data)}\n")
            
            # Write airport data
            for code, lat, lon in airports_data:
                f.write(f"{code}\t{lat}\t{lon}\n")
            
            # Write number of constraints
            f.write(f"{len(constraints_data)}\n")
            
            # Write constraints
            for beg, end in constraints_data:
                f.write(f"{beg}\t{end}\n")
            
            # Write start airport (if provided)
            if start_airport:
                f.write(f"{start_airport}\n")
        
        print(f"✓ Successfully created {output_file} from Excel")
        print(f"  - {len(airports_data)} airports")
        print(f"  - {len(constraints_data)} constraints")
        if start_airport:
            print(f"  - Start airport: {start_airport}")
        return True
    
    except Exception as e:
        print(f"Error writing output file: {e}")
        return False


if __name__ == "__main__":
    excel_file = "C:/Users/ericc/OneDrive/FlightPlans/Flight Simulator X Files/Flight Tools-EC2021.xlsm"  # Change this to your Excel filename
    output_file = "input.txt"
    
    if excel_to_input(excel_file, output_file):
        print("\nYou can now run: FP.exe")
    else:
        print("\nConversion failed. Check your Excel file and named ranges.")
